# -*- coding: utf-8 -*-
# @Time    : 2022/6/8 17:30
# @Author  : Tom_zc
# @FileName: collect_elastic_public_ip_by_yaml.py
# @Software: PyCharm
import os

import requests
import argparse
import time
import openpyxl
import yaml
from abc import abstractmethod
from functools import wraps

from requests.packages.urllib3.exceptions import InsecureRequestWarning
from huaweicloudsdkcore.auth.credentials import BasicCredentials
from huaweicloudsdkcore.http.http_config import HttpConfig
from huaweicloudsdkcore.client import Client
from huaweicloudsdkeip.v2 import EipClient as EipClientV2
from huaweicloudsdkeip.v2 import ListPublicipsRequest as ListPublicipsRequestV2
from huaweicloudsdkeip.v3 import EipClient as EipClientV3
from huaweicloudsdkeip.v3 import ListPublicipsRequest as ListPublicipsRequestV3
from huaweicloudsdknat.v2 import NatClient, ListNatGatewaysRequest
from huaweicloudsdkelb.v2 import ElbClient, ListLoadbalancersRequest
from huaweicloudsdkbms.v1 import BmsClient, ListBareMetalServersRequest
from huaweicloudsdkecs.v2 import EcsClient, NovaListServersDetailsRequest
from huaweicloudsdkrds.v3 import RdsClient, ListInstancesRequest

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)


# noinspection DuplicatedCode
class GlobalConfig(object):
    base_path = os.path.dirname(__file__)
    excel_path = os.path.join(base_path, "公网IP统计表.xlsx")
    config_path = os.path.join(base_path, "collect_elastic_public_ip.yaml")
    excel_title = ["弹性公网IP", "IPv6地址", "弹性公网IP ID", "状态", "类型", "带宽名称", "带宽ID", "带宽大小(Mbit/s)",
                   "实例类型", "实例名称", "实例ID", "实例归属区域", "创建时间"]
    zone_alias_dict = {
        "cn-north-1": "华北-北京一",
        "cn-north-4": "华北-北京四",
        "cn-north-5": "华北-乌兰察布二零一",
        "cn-north-6": "华北-乌兰察布二零二",
        "cn-north-9": "华北-乌兰察布一",
        "cn-east-3": "华东-上海一",
        "cn-east-2": "华东-上海二",
        "cn-south-1": "华南-广州",
        "cn-south-4": "华南-广州-友好用户环境",
        "cn-southwest-2": "西南-贵阳一",
        "ap-southeast-1": "中国-香港",
        "ap-southeast-2": "亚太-曼谷",
        "ap-southeast-3": "亚太-新加坡",
        "af-south-1": "非洲-约翰内斯堡",
        "na-mexico-1": "拉美-墨西哥城一",
        "la-north-2": "拉美-墨西哥城二",
        "sa-brazil-1": "拉美-圣保罗一",
        "la-south-2": "拉美-圣地亚哥",
        "ru-northwest-2": "俄罗斯-莫斯科二",
    }
    need_delete_sheet_name = "Sheet"
    eip_v2_zone = ["cn-south-4", ]


# noinspection DuplicatedCode
class EndPoint(object):
    vpc_endpoint = "https://vpc.{}.myhuaweicloud.com"
    nat_endpoint = "https://nat.{}.myhuaweicloud.com"
    elb_endpoint = "https://elb.{}.myhuaweicloud.com"
    bms_endpoint = "https://bms.{}.myhuaweicloud.com"
    ecs_endpoint = "https://ecs.{}.myhuaweicloud.com"
    rds_endpoint = "https://rds.{}.myhuaweicloud.com"


class IpStatus(object):
    FREEZED = "冻结"
    BIND_ERROR = "绑定失败"
    BINDING = "绑定中"
    PENDING_CREATE = "创建中"
    PENDING_DELETE = "释放中"
    NOTIFYING = "绑定中"
    NOTIFY_DELETE = "释放中"
    PENDING_UPDATE = "更新中"
    DOWN = "未绑定"
    ACTIVE = "绑定"
    ELB = "绑定ELB"
    VPN = "绑定VPN"
    ERROR = "失败"


class IpType(object):
    EIP = "全动态BGP"


def func_retry(tries=3, delay=1):
    def deco_retry(fn):
        @wraps(fn)
        def inner(*args, **kwargs):
            for i in range(tries):
                try:
                    return fn(*args, **kwargs)
                except Exception as e:
                    print(e)
                    time.sleep(delay)
            else:
                print("func_retry: {} failed".format(fn.__name__))

        return inner

    return deco_retry


# noinspection PyUnresolvedReferences,DuplicatedCode
class BaseInstance(object):
    def __init__(self, base_client, config, credentials, endpoint):
        if not issubclass(base_client, Client):
            raise Exception("base client must be client")
        self.base_client = base_client.new_builder() \
            .with_http_config(config) \
            .with_credentials(credentials) \
            .with_endpoint(endpoint) \
            .build()

    @abstractmethod
    def set_req_method(self):
        pass

    def parse_response_data(self, response_dict):
        dict_data = dict()
        for item in response_dict[self.ret_filed]:
            dict_data[item["id"]] = {
                "name": item["name"],
                "instance_type": self.instance_name
            }
        return dict_data

    def show_infos(self, *args, **kwargs):
        info_request, method = self.set_req_method()
        show_infos_req = info_request(*args, **kwargs)
        show_infos_method = getattr(self.base_client, method)
        ret = show_infos_method(show_infos_req)
        return ret.to_dict()


class EipInstanceV2(BaseInstance):
    def __init__(self, *args, **kwargs):
        super(EipInstanceV2, self).__init__(*args, **kwargs)

    def set_req_method(self):
        return ListPublicipsRequestV2, "list_publicips"

    def parse_response_data(self, response_dict):
        return response_dict['publicips']


class EipInstanceV3(BaseInstance):
    def __init__(self, *args, **kwargs):
        super(EipInstanceV3, self).__init__(*args, **kwargs)

    def set_req_method(self):
        return ListPublicipsRequestV3, "list_publicips"

    def parse_response_data(self, response_dict):
        return response_dict['publicips']


class NatInstance(BaseInstance):
    def __init__(self, *args, **kwargs):
        super(NatInstance, self).__init__(*args, **kwargs)
        self.instance_name = "NAT网关"
        self.ret_filed = "nat_gateways"

    def set_req_method(self):
        return ListNatGatewaysRequest, "list_nat_gateways"


class LoadBalanceInstance(BaseInstance):
    def __init__(self, *args, **kwargs):
        super(LoadBalanceInstance, self).__init__(*args, **kwargs)
        self.instance_name = "负载均衡器"
        self.ret_filed = "loadbalancers"

    def set_req_method(self):
        return ListLoadbalancersRequest, "list_loadbalancers"


class BMSInstance(BaseInstance):
    def __init__(self, *args, **kwargs):
        super(BMSInstance, self).__init__(*args, **kwargs)
        self.instance_name = "裸金属服务器"
        self.ret_filed = "servers"

    def set_req_method(self):
        return ListBareMetalServersRequest, "list_bare_metal_servers"


class EcsInstance(BaseInstance):
    def __init__(self, *args, **kwargs):
        super(EcsInstance, self).__init__(*args, **kwargs)
        self.instance_name = "云服务器"
        self.ret_filed = "servers"

    def set_req_method(self):
        return NovaListServersDetailsRequest, "nova_list_servers_details"


class RdsInstance(BaseInstance):
    def __init__(self, *args, **kwargs):
        super(RdsInstance, self).__init__(*args, **kwargs)
        self.instance_name = "云数据库 RDS"
        self.ret_filed = "instances"

    def set_req_method(self):
        return ListInstancesRequest, "list_instances"


# noinspection DuplicatedCode
class EipTools(object):
    def __init__(self, *args, **kwargs):
        super(EipTools, self).__init__(*args, **kwargs)

    @classmethod
    def get_eip_config(cls):
        config = HttpConfig.get_default_config()
        config.ignore_ssl_verification = True
        return config

    @classmethod
    def output_excel(cls, eip_info_list, username):
        if os.path.exists(GlobalConfig.excel_path):
            work_book = openpyxl.load_workbook(GlobalConfig.excel_path)
        else:
            work_book = openpyxl.Workbook()
        if username not in work_book.get_sheet_names():
            work_book.create_sheet(username)
        if GlobalConfig.need_delete_sheet_name in work_book.get_sheet_names():
            need_remove_sheet = work_book.get_sheet_by_name(GlobalConfig.need_delete_sheet_name)
            work_book.remove_sheet(need_remove_sheet)
        table = work_book.get_sheet_by_name(username)
        table.delete_rows(1, 65536)
        table.append(GlobalConfig.excel_title)
        for eip_info in eip_info_list:
            table.append(eip_info)
        work_book.save(GlobalConfig.excel_path)

    @classmethod
    def get_device_info(cls, instance_list):
        ret_dict = dict()
        for instance_temp in instance_list:
            instance_info = instance_temp.show_infos()
            device_info = instance_temp.parse_response_data(instance_info)
            for key, value in device_info.items():
                if key not in ret_dict.keys():
                    ret_dict[key] = value
        return ret_dict

    @classmethod
    def parse_input_args(cls):
        par = argparse.ArgumentParser()
        par.add_argument("-config_path", "--config_path", help="The config path of object", required=False)
        args = par.parse_args()
        return args

    @staticmethod
    def load_yaml(file_path, method="load"):
        """
        method: load_all/load
        """
        yaml_load_method = getattr(yaml, method)
        with open(file_path, "r", encoding="utf-8") as file:
            content = yaml_load_method(file, Loader=yaml.FullLoader)
        return content

    @classmethod
    def check_config_data(cls, config_list):
        for config_temp in config_list:
            if config_temp.get("account") is None:
                raise Exception("Account is invalid")
            if config_temp.get("ak") is None:
                raise Exception("Ak is invalid")
            if config_temp.get("sk") is None:
                raise Exception("Sk is invalid")
            if not isinstance(config_temp.get("project_info"), list):
                raise Exception("project info must be dict")
            for project_temp in config_temp["project_info"]:
                if project_temp.get("project_id") is None:
                    raise Exception("project_id is invalid")
                if project_temp.get("zone") is None:
                    raise Exception("project_id is invalid")

    @classmethod
    def parse_ips_v2(cls, eip_list, zone):
        result_list = list()
        zone_alias = GlobalConfig.zone_alias_dict.get(zone, zone)
        for eip_info in eip_list:
            temp = list()
            temp.append(eip_info['public_ip_address'])
            temp.append(eip_info['public_ipv6_address'])
            temp.append(eip_info['id'])
            temp.append(IpStatus.__dict__.get(eip_info['status'], "未知中"))
            temp.append(IpType.__dict__.get(eip_info['type'], eip_info['type']))
            temp.append(eip_info["bandwidth_name"])
            temp.append(eip_info["bandwidth_id"])
            temp.append(eip_info["bandwidth_size"])
            temp.append(None)
            temp.append(None)
            temp.append(None)
            temp.append(zone_alias)
            temp.append(str(eip_info["create_time"]))
            result_list.append(temp)
        return result_list

    @classmethod
    def parse_ips_v3(cls, eip_list, device_info_dict, zone):
        result_list = list()
        zone_alias = GlobalConfig.zone_alias_dict.get(zone, zone)
        for eip_info in eip_list:
            temp = list()
            temp.append(eip_info['public_ip_address'])
            temp.append(eip_info['public_ipv6_address'])
            temp.append(eip_info['id'])
            temp.append(IpStatus.__dict__.get(eip_info['status'], "未知中"))
            temp.append(IpType.__dict__.get(eip_info['type'], "未知中"))
            temp.append(eip_info['bandwidth']["name"])
            temp.append(eip_info['bandwidth']["id"])
            temp.append(eip_info['bandwidth']["size"])
            # 1.如果ip绑定的是负载均衡,则vnic为空， device_id为associate_instance_id
            # 2.如果ip绑定的是RDS，则device_id为空并且instance_type为RDS， device_id为instance_id
            # 3.如果ip绑定的是PORT, 则很有可能为虚拟ip.
            if not isinstance(eip_info['vnic'], dict):
                device_id = eip_info['associate_instance_id']
                if device_id and device_id in device_info_dict.keys():
                    temp.append(device_info_dict[device_id]["instance_type"])
                    temp.append(device_info_dict[device_id]["name"])
                    temp.append(device_id)
                elif eip_info["status"] != IpStatus.ACTIVE:
                    temp.append(None)
                    temp.append(None)
                    temp.append(None)
                else:
                    raise Exception("script need to update， reason:1!!!")
            elif not eip_info["vnic"]["device_id"] and eip_info["vnic"]["instance_type"] == "RDS":
                device_id = eip_info["vnic"]["instance_id"]
                if device_id and device_id in device_info_dict.keys():
                    temp.append(device_info_dict[device_id]["instance_type"])
                    temp.append(device_info_dict[device_id]["name"])
                    temp.append(device_id)
                else:
                    raise Exception("script need to update， reason:2!!!")
            elif eip_info["vnic"]["device_id"] and eip_info["vnic"]["device_id"] in device_info_dict.keys():
                temp.append(device_info_dict[eip_info["vnic"]["device_id"]]["instance_type"])
                temp.append(device_info_dict[eip_info["vnic"]["device_id"]]["name"])
                temp.append(eip_info["vnic"]["device_id"])
            elif eip_info["associate_instance_type"] == "PORT":
                temp.append("虚拟IP地址")
                temp.append(eip_info["vnic"]["private_ip_address"])
                temp.append(eip_info["associate_instance_id"])
            else:
                print("eip_info:{}".format(eip_info))
                raise Exception("script need to update， reason:3!!!")
            temp.append(zone_alias)
            temp.append(str(eip_info["created_at"]))
            result_list.append(temp)
        return result_list

    @func_retry()
    def get_data_list(self, eip_tools, project_temp, ak, sk):
        project_id = project_temp["project_id"]
        zone = project_temp["zone"]
        config = eip_tools.get_eip_config()
        credentials = BasicCredentials(ak, sk, project_id)
        if zone in GlobalConfig.eip_v2_zone:
            eip_instance = EipInstanceV2(EipClientV2, config, credentials, EndPoint.vpc_endpoint.format(zone))
        else:
            eip_instance = EipInstanceV3(EipClientV3, config, credentials, EndPoint.vpc_endpoint.format(zone))
        nat_instance = NatInstance(NatClient, config, credentials, EndPoint.nat_endpoint.format(zone))
        elb_instance = LoadBalanceInstance(ElbClient, config, credentials, EndPoint.elb_endpoint.format(zone))
        bms_instance = BMSInstance(BmsClient, config, credentials, EndPoint.bms_endpoint.format(zone))
        ecs_instance = EcsInstance(EcsClient, config, credentials, EndPoint.ecs_endpoint.format(zone))
        rds_instance = RdsInstance(RdsClient, config, credentials, EndPoint.rds_endpoint.format(zone))
        query_device_lists = [nat_instance, elb_instance, bms_instance, ecs_instance, rds_instance]
        device_info_dict = eip_tools.get_device_info(query_device_lists)
        eip_dict = eip_instance.show_infos()
        eip_list = eip_instance.parse_response_data(eip_dict)
        if zone in GlobalConfig.eip_v2_zone:
            result_list = eip_tools.parse_ips_v2(eip_list, zone)
        else:
            result_list = eip_tools.parse_ips_v3(eip_list, device_info_dict, zone)
        return result_list


# noinspection DuplicatedCode
def main():
    eip_tools = EipTools()
    input_args = eip_tools.parse_input_args()
    print("##################1.start to parse params #############")
    if not input_args.config_path:
        config_path = GlobalConfig.config_path
    else:
        config_path = input_args.config_path
    config_list = eip_tools.load_yaml(config_path)
    eip_tools.check_config_data(config_list)
    print("############2.start to collect and output to excel######")
    for config_item in config_list:
        result_list = list()
        username = config_item['account']
        ak = config_item["ak"]
        sk = config_item["sk"]
        project_info = config_item["project_info"]
        print("Collect the username of info:{}".format(username))
        for project_temp in project_info:
            print("Collect the zone of info:{}".format(project_temp["zone"]))
            ret_temp = eip_tools.get_data_list(eip_tools, project_temp, ak, sk)
            result_list.extend(ret_temp or [])
        print("Write the data to excel...")
        if result_list:
            eip_tools.output_excel(result_list, username)
        else:
            print("There is no data to write to excel.")
    print("##################3.finish################")


if __name__ == "__main__":
    main()
